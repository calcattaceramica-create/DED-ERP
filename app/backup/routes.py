from flask import render_template, redirect, url_for, flash, request, send_file, current_app, make_response
from flask_login import login_required, current_user
from app.backup import bp
from app.auth.decorators import permission_required
from app import db
import os
import subprocess
from datetime import datetime
import zipfile
import json
import io

@bp.route('/')
@login_required
@permission_required('settings.manage')
def index():
    """Backup management page"""
    backup_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'backups')
    
    # Create backup directory if it doesn't exist
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
    
    # Get list of existing backups
    backups = []
    if os.path.exists(backup_dir):
        for filename in os.listdir(backup_dir):
            if filename.endswith('.zip'):
                filepath = os.path.join(backup_dir, filename)
                file_stat = os.stat(filepath)
                backups.append({
                    'filename': filename,
                    'size': file_stat.st_size,
                    'size_mb': round(file_stat.st_size / (1024 * 1024), 2),
                    'created_at': datetime.fromtimestamp(file_stat.st_ctime),
                    'path': filepath
                })
    
    # Sort by creation date (newest first)
    backups.sort(key=lambda x: x['created_at'], reverse=True)
    
    # Get backup settings
    settings = get_backup_settings()
    
    return render_template('backup/index.html', backups=backups, settings=settings)

@bp.route('/create', methods=['POST'])
@login_required
@permission_required('settings.manage')
def create_backup():
    """Create a new backup (PostgreSQL pg_dump + uploads)"""
    try:
        basedir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        backup_dir = os.path.join(basedir, 'backups')

        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'backup_{timestamp}.zip'
        backup_path = os.path.join(backup_dir, backup_filename)
        sql_dump_path = os.path.join(backup_dir, f'db_{timestamp}.sql')

        # --- PostgreSQL dump ---
        db_url = current_app.config.get('SQLALCHEMY_DATABASE_URI', '')
        pg_dump_ok = False
        if 'postgresql' in db_url:
            try:
                import urllib.parse
                parsed = urllib.parse.urlparse(db_url)
                pg_env = os.environ.copy()
                pg_env['PGPASSWORD'] = parsed.password or ''

                # Try full path first (Windows), then fallback to PATH
                pg_dump_candidates = [
                    r'C:\Program Files\PostgreSQL\16\bin\pg_dump.exe',
                    r'C:\Program Files\PostgreSQL\15\bin\pg_dump.exe',
                    r'C:\Program Files\PostgreSQL\17\bin\pg_dump.exe',
                    'pg_dump',
                ]
                pg_dump_exe = None
                for candidate in pg_dump_candidates:
                    if os.path.exists(candidate) or candidate == 'pg_dump':
                        pg_dump_exe = candidate
                        break

                if pg_dump_exe:
                    pg_dump_cmd = [
                        pg_dump_exe,
                        '-h', parsed.hostname or 'localhost',
                        '-p', str(parsed.port or 5432),
                        '-U', parsed.username or 'postgres',
                        '-d', parsed.path.lstrip('/'),
                        '-f', sql_dump_path,
                        '--no-password'
                    ]
                    result = subprocess.run(pg_dump_cmd, env=pg_env, capture_output=True, timeout=120)
                    if result.returncode == 0:
                        pg_dump_ok = True
            except Exception as pg_err:
                pg_dump_ok = False

        with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Add SQL dump
            if pg_dump_ok and os.path.exists(sql_dump_path):
                zipf.write(sql_dump_path, 'database_dump.sql')
                os.remove(sql_dump_path)

            # Backup uploads folder
            uploads_dir = os.path.join(basedir, 'uploads')
            if os.path.exists(uploads_dir):
                for root, dirs, files in os.walk(uploads_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, basedir)
                        zipf.write(file_path, arcname)

            # Backup config file
            config_path = os.path.join(basedir, 'config.py')
            if os.path.exists(config_path):
                zipf.write(config_path, 'config.py')

            backup_info = {
                'created_at': datetime.now().isoformat(),
                'created_by': current_user.username,
                'database': 'PostgreSQL' if pg_dump_ok else 'skipped',
                'version': '2.0'
            }
            zipf.writestr('backup_info.json', json.dumps(backup_info, indent=2, ensure_ascii=False))

        if pg_dump_ok:
            flash(f'تم انشاء النسخة الاحتياطية بنجاح (مع قاعدة البيانات): {backup_filename}', 'success')
        else:
            flash(f'تم انشاء النسخة الاحتياطية (ملفات فقط، pg_dump غير متاح): {backup_filename}', 'warning')

    except Exception as e:
        flash(f'حدث خطأ اثناء انشاء النسخة الاحتياطية: {str(e)}', 'danger')

    return redirect(url_for('backup.index'))

@bp.route('/download/<filename>')
@login_required
@permission_required('settings.manage')
def download_backup(filename):
    """Download a backup file"""
    try:
        backup_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'backups')
        backup_path = os.path.join(backup_dir, filename)
        
        if not os.path.exists(backup_path):
            flash('الملف غير موجود', 'danger')
            return redirect(url_for('backup.index'))

        return send_file(backup_path, as_attachment=True, download_name=filename)

    except Exception as e:
        flash(f'حدث خطا اثناء تحميل النسخة الاحتياطية: {str(e)}', 'danger')
        return redirect(url_for('backup.index'))

@bp.route('/delete/<filename>', methods=['POST'])
@login_required
@permission_required('settings.manage')
def delete_backup(filename):
    """Delete a backup file"""
    try:
        backup_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'backups')
        backup_path = os.path.join(backup_dir, filename)
        
        if os.path.exists(backup_path):
            os.remove(backup_path)
            flash(f'تم حذف النسخة الاحتياطية: {filename}', 'success')
        else:
            flash('الملف غير موجود', 'danger')

    except Exception as e:
        flash(f'حدث خطا اثناء حذف النسخة الاحتياطية: {str(e)}', 'danger')
    
    return redirect(url_for('backup.index'))

@bp.route('/export-excel')
@login_required
@permission_required('settings.manage')
def export_excel():
    """Export backup list as Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        backup_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'backups')
        backups = []
        if os.path.exists(backup_dir):
            for filename in os.listdir(backup_dir):
                if filename.endswith('.zip'):
                    filepath = os.path.join(backup_dir, filename)
                    file_stat = os.stat(filepath)
                    backups.append({
                        'filename': filename,
                        'size_mb': round(file_stat.st_size / (1024 * 1024), 2),
                        'created_at': datetime.fromtimestamp(file_stat.st_ctime),
                    })
        backups.sort(key=lambda x: x['created_at'], reverse=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Backups'

        # Title row
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = 'Backup Management Report'
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill('solid', fgColor='1a1a2e')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Sub title
        ws.merge_cells('A2:D2')
        sub_cell = ws['A2']
        sub_cell.value = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} | Total: {len(backups)} backup(s)'
        sub_cell.font = Font(size=10, color='666666')
        sub_cell.alignment = Alignment(horizontal='center')

        # Header row
        headers = ['#', 'File Name', 'Size (MB)', 'Created At']
        header_fill = PatternFill('solid', fgColor='2d6a4f')
        thin = Side(style='thin', color='cccccc')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[3].height = 22

        # Data rows
        for i, b in enumerate(backups, 1):
            row = i + 3
            fill = PatternFill('solid', fgColor='f5f5f5') if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
            values = [i, b['filename'], b['size_mb'], b['created_at'].strftime('%Y-%m-%d %H:%M:%S')]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = fill
                cell.border = border

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 22

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename_excel = f'backups_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = make_response(buf.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename_excel}'
        return response

    except Exception as e:
        flash(f'خطأ في إنشاء Excel: {str(e)}', 'danger')
        return redirect(url_for('backup.index'))


@bp.route('/export-pdf')
@login_required
@permission_required('settings.manage')
def export_pdf():
    """Export backup list as PDF"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import arabic_reshaper
        from bidi.algorithm import get_display

        def ar(text):
            try:
                reshaped = arabic_reshaper.reshape(str(text))
                return get_display(reshaped)
            except Exception:
                return str(text)

        # Get backup list
        backup_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'backups')
        backups = []
        if os.path.exists(backup_dir):
            for filename in os.listdir(backup_dir):
                if filename.endswith('.zip'):
                    filepath = os.path.join(backup_dir, filename)
                    file_stat = os.stat(filepath)
                    backups.append({
                        'filename': filename,
                        'size_mb': round(file_stat.st_size / (1024 * 1024), 2),
                        'created_at': datetime.fromtimestamp(file_stat.st_ctime),
                    })
        backups.sort(key=lambda x: x['created_at'], reverse=True)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                rightMargin=1.5*cm, leftMargin=1.5*cm,
                                topMargin=2*cm, bottomMargin=2*cm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=16,
                                     alignment=1, spaceAfter=6)
        sub_style = ParagraphStyle('sub', fontName='Helvetica', fontSize=10,
                                   alignment=1, spaceAfter=12, textColor=colors.grey)

        elements = []
        elements.append(Paragraph('Backup Management Report', title_style))
        elements.append(Paragraph(
            f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} | Total: {len(backups)} backup(s)',
            sub_style))
        elements.append(Spacer(1, 0.4*cm))

        # Table
        header = ['#', ar('اسم الملف'), ar('الحجم'), ar('تاريخ الإنشاء')]
        rows = [header]
        for i, b in enumerate(backups, 1):
            rows.append([
                str(i),
                ar(b['filename']),
                f"{b['size_mb']} MB",
                b['created_at'].strftime('%Y-%m-%d %H:%M'),
            ])

        col_widths = [1*cm, 9*cm, 3*cm, 5*cm]
        t = Table(rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)

        doc.build(elements)
        buf.seek(0)

        filename_pdf = f'backups_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = make_response(buf.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={filename_pdf}'
        return response

    except Exception as e:
        flash(f'خطأ في إنشاء PDF: {str(e)}', 'danger')
        return redirect(url_for('backup.index'))


def get_backup_settings():
    """Get backup settings from database or return defaults"""
    from app.models_settings import SystemSettings
    
    auto_backup = SystemSettings.query.filter_by(setting_key='auto_backup_enabled').first()
    backup_frequency = SystemSettings.query.filter_by(setting_key='backup_frequency').first()
    
    return {
        'auto_backup_enabled': auto_backup.setting_value == 'true' if auto_backup else False,
        'backup_frequency': backup_frequency.setting_value if backup_frequency else 'daily'
    }

