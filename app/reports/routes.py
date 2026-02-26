from flask import render_template, redirect, url_for, request, current_app, send_file
from flask_login import login_required
from app.auth.decorators import permission_required
from app.reports import bp
from app import db
from app.models import *
from sqlalchemy import func
from datetime import datetime, timedelta
import io

@bp.route('/')
@login_required
@permission_required('reports.view')
def index():
    """Reports dashboard"""
    return render_template('reports/index.html')

@bp.route('/sales')
@login_required
@permission_required('reports.sales')
def sales_report():
    """Sales report"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = SalesInvoice.query.filter_by(status='confirmed')

    if start_date:
        query = query.filter(SalesInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(SalesInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())

    invoices = query.all()

    total_sales = sum(inv.total_amount for inv in invoices)
    total_tax = sum(inv.tax_amount for inv in invoices)

    # Get currency settings
    from app.models import Company
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'EUR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'Euro')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    return render_template('reports/sales.html',
                         invoices=invoices,
                         total_sales=total_sales,
                         total_tax=total_tax,
                         start_date=start_date,
                         end_date=end_date,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)

@bp.route('/purchases')
@login_required
@permission_required('reports.purchases')
def purchases_report():
    """Purchases report"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = PurchaseInvoice.query.filter(PurchaseInvoice.status != 'cancelled')

    if start_date:
        query = query.filter(PurchaseInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(PurchaseInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())

    invoices = query.all()

    total_purchases = sum(inv.total_amount for inv in invoices)
    total_tax = sum(inv.tax_amount for inv in invoices)

    # Get currency settings
    currency_code = current_app.config.get('CURRENCY', 'EUR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'Euro')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    # Debug
    print(f"DEBUG: invoices count = {len(invoices)}")
    print(f"DEBUG: total_purchases = {total_purchases}")
    print(f"DEBUG: total_tax = {total_tax}")
    print(f"DEBUG: currency_name = {currency_name}")

    return render_template('reports/purchases.html',
                         invoices=invoices,
                         total_purchases=total_purchases,
                         total_tax=total_tax,
                         start_date=start_date,
                         end_date=end_date,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)

@bp.route('/inventory')
@login_required
@permission_required('reports.inventory')
def inventory_report():
    """Inventory report"""
    products = Product.query.filter_by(is_active=True, track_inventory=True).all()

    inventory_data = []
    for product in products:
        stock_qty = product.get_stock()
        inventory_data.append({
            'product': product,
            'stock': stock_qty,
            'value': stock_qty * product.cost_price
        })

    total_value = sum(item['value'] for item in inventory_data)

    # Get company settings for currency
    from app.models import Company
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'SAR')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', 'ر.س')

    return render_template('reports/inventory.html',
                         inventory_data=inventory_data,
                         total_value=total_value,
                         currency_code=currency_code,
                         currency_symbol=currency_symbol)

@bp.route('/profit-loss')
@login_required
@permission_required('reports.financial')
def profit_loss():
    """Profit and Loss statement"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Calculate revenue
    revenue_query = db.session.query(func.sum(SalesInvoice.total_amount)).filter(SalesInvoice.status != 'cancelled')
    if start_date:
        revenue_query = revenue_query.filter(SalesInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        revenue_query = revenue_query.filter(SalesInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())

    total_revenue = revenue_query.scalar() or 0

    # Calculate cost of goods sold (COGS) - from sales invoice items
    sales_query = SalesInvoice.query.filter(SalesInvoice.status != 'cancelled')
    if start_date:
        sales_query = sales_query.filter(SalesInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        sales_query = sales_query.filter(SalesInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())

    sales_invoices = sales_query.all()

    total_cogs = 0
    for invoice in sales_invoices:
        for item in invoice.items:
            total_cogs += item.product.cost_price * item.quantity

    gross_profit = total_revenue - total_cogs

    # Get currency settings
    currency_code = current_app.config.get('CURRENCY', 'EUR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'Euro')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    return render_template('reports/profit_loss.html',
                         total_revenue=total_revenue,
                         total_cogs=total_cogs,
                         gross_profit=gross_profit,
                         start_date=start_date,
                         end_date=end_date,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)

@bp.route('/low-stock')
@login_required
@permission_required('reports.inventory')
def low_stock_report():
    """Low stock products report"""
    products = Product.query.filter_by(is_active=True, track_inventory=True).all()

    low_stock_products = []
    for product in products:
        stock_qty = product.get_stock()
        if stock_qty <= product.min_stock:
            low_stock_products.append({
                'product': product,
                'current_stock': stock_qty,
                'min_stock': product.min_stock,
                'shortage': product.min_stock - stock_qty
            })

    return render_template('reports/low_stock.html',
                         low_stock_products=low_stock_products)

@bp.route('/stock-movement')
@login_required
@permission_required('reports.inventory')
def stock_movement_report():
    """Stock movement report"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    product_id = request.args.get('product_id', type=int)
    warehouse_id = request.args.get('warehouse_id', type=int)

    query = StockMovement.query

    if start_date:
        query = query.filter(StockMovement.created_at >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(StockMovement.created_at <= datetime.strptime(end_date + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    if product_id:
        query = query.filter_by(product_id=product_id)
    if warehouse_id:
        query = query.filter_by(warehouse_id=warehouse_id)

    movements = query.order_by(StockMovement.created_at.desc()).all()

    products = Product.query.filter_by(is_active=True).order_by(Product.name).all()
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()

    return render_template('reports/stock_movement.html',
                         movements=movements,
                         products=products,
                         warehouses=warehouses,
                         start_date=start_date,
                         end_date=end_date,
                         selected_product_id=product_id,
                         selected_warehouse_id=warehouse_id)

@bp.route('/sales-by-product')
@login_required
@permission_required('reports.sales')
def sales_by_product():
    """Sales report by product"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = db.session.query(
        Product.name,
        Product.code,
        func.sum(SalesInvoiceItem.quantity).label('total_qty'),
        func.sum(SalesInvoiceItem.total).label('total_amount')
    ).join(SalesInvoiceItem).join(SalesInvoice).filter(
        SalesInvoice.status != 'cancelled'
    )

    if start_date:
        query = query.filter(SalesInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(SalesInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())

    results = query.group_by(Product.id).order_by(func.sum(SalesInvoiceItem.total).desc()).all()

    total_qty = sum(r.total_qty for r in results)
    total_amount = sum(r.total_amount for r in results)

    return render_template('reports/sales_by_product.html',
                         results=results,
                         total_qty=total_qty,
                         total_amount=total_amount,
                         start_date=start_date,
                         end_date=end_date)

@bp.route('/sales-by-customer')
@login_required
@permission_required('reports.sales')
def sales_by_customer():
    """Sales report by customer"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = db.session.query(
        Customer.name,
        Customer.code,
        func.count(SalesInvoice.id).label('invoice_count'),
        func.sum(SalesInvoice.total_amount).label('total_amount')
    ).join(SalesInvoice).filter(
        SalesInvoice.status != 'cancelled'
    )

    if start_date:
        query = query.filter(SalesInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(SalesInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())

    results = query.group_by(Customer.id).order_by(func.sum(SalesInvoice.total_amount).desc()).all()

    total_invoices = sum(r.invoice_count for r in results)
    total_amount = sum(r.total_amount for r in results)

    return render_template('reports/sales_by_customer.html',
                         results=results,
                         total_invoices=total_invoices,
                         total_amount=total_amount,
                         start_date=start_date,
                         end_date=end_date)

@bp.route('/sales-monthly')
@login_required
@permission_required('reports.sales')
def sales_monthly():
    """Monthly sales report"""
    year = request.args.get('year', type=int)
    if not year:
        year = datetime.now().year

    start_date = datetime(year, 1, 1).date()
    end_date = datetime(year, 12, 31).date()

    invoices = SalesInvoice.query.filter(
        SalesInvoice.status != 'cancelled',
        SalesInvoice.invoice_date >= start_date,
        SalesInvoice.invoice_date <= end_date
    ).all()

    # Group by month
    monthly_dict = {}
    for invoice in invoices:
        month_num = invoice.invoice_date.month
        if month_num not in monthly_dict:
            monthly_dict[month_num] = {'invoice_count': 0, 'total_amount': 0, 'total_tax': 0}
        monthly_dict[month_num]['invoice_count'] += 1
        monthly_dict[month_num]['total_amount'] += invoice.total_amount
        monthly_dict[month_num]['total_tax'] += (invoice.tax_amount or 0)

    month_names = [
        'يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
        'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'
    ]

    months_data = []
    for month_num in range(1, 13):
        data = monthly_dict.get(month_num, {'invoice_count': 0, 'total_amount': 0, 'total_tax': 0})
        months_data.append({
            'month_num': month_num,
            'month_name': month_names[month_num - 1],
            'invoice_count': data['invoice_count'],
            'total_amount': data['total_amount'],
            'total_tax': data['total_tax']
        })

    total_invoices = sum(m['invoice_count'] for m in months_data)
    total_sales = sum(m['total_amount'] for m in months_data)
    total_tax = sum(m['total_tax'] for m in months_data)

    from app.models import Company
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'EUR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'Euro')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    # Get available years
    all_invoices = SalesInvoice.query.filter(SalesInvoice.status != 'cancelled').all()
    years_set = {inv.invoice_date.year for inv in all_invoices if inv.invoice_date}
    available_years = sorted(list(years_set), reverse=True) or [datetime.now().year]

    return render_template('reports/sales_monthly.html',
                         months_data=months_data,
                         total_invoices=total_invoices,
                         total_sales=total_sales,
                         total_tax=total_tax,
                         year=year,
                         available_years=available_years,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)

@bp.route('/purchases-by-product')
@login_required
@permission_required('reports.purchases')
def purchases_by_product():
    """Purchases report by product"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    product_id = request.args.get('product_id', type=int)

    # Base query for purchase invoice items
    query = db.session.query(
        Product.id,
        Product.name,
        Product.code,
        func.sum(PurchaseInvoiceItem.quantity).label('total_quantity'),
        func.sum(PurchaseInvoiceItem.total).label('total_amount'),
        func.count(func.distinct(PurchaseInvoiceItem.invoice_id)).label('invoice_count')
    ).join(
        PurchaseInvoiceItem, Product.id == PurchaseInvoiceItem.product_id
    ).join(
        PurchaseInvoice, PurchaseInvoiceItem.invoice_id == PurchaseInvoice.id
    ).filter(
        PurchaseInvoice.status != 'cancelled'
    )

    # Apply filters
    if start_date:
        query = query.filter(PurchaseInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(PurchaseInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if product_id:
        query = query.filter(Product.id == product_id)

    # Group by product
    query = query.group_by(Product.id, Product.name, Product.code)

    # Order by total amount descending
    query = query.order_by(func.sum(PurchaseInvoiceItem.total).desc())

    products_data = query.all()

    # Calculate totals
    total_quantity = sum(p.total_quantity or 0 for p in products_data)
    total_amount = sum(p.total_amount or 0 for p in products_data)
    total_invoices = sum(p.invoice_count or 0 for p in products_data)

    # Get all products for filter dropdown
    all_products = Product.query.filter_by(is_active=True, is_purchasable=True).order_by(Product.name).all()

    # Get currency info
    currencies = current_app.config.get('CURRENCIES', {})
    default_currency = current_app.config.get('DEFAULT_CURRENCY', 'EUR')
    currency_info = currencies.get(default_currency, {})
    currency_code = default_currency
    currency_name = currency_info.get('name', 'Euro')
    currency_symbol = currency_info.get('symbol', '€')

    return render_template('reports/purchases_by_product.html',
                         products_data=products_data,
                         total_quantity=total_quantity,
                         total_amount=total_amount,
                         total_invoices=total_invoices,
                         all_products=all_products,
                         start_date=start_date,
                         end_date=end_date,
                         selected_product_id=product_id,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)

@bp.route('/purchases-by-supplier')
@login_required
@permission_required('reports.purchases')
def purchases_by_supplier():
    """Purchases report grouped by supplier"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    supplier_id = request.args.get('supplier_id', type=int)

    query = db.session.query(
        Supplier.id,
        Supplier.name,
        Supplier.code,
        func.count(func.distinct(PurchaseInvoice.id)).label('invoice_count'),
        func.sum(PurchaseInvoice.total_amount).label('total_amount'),
        func.sum(PurchaseInvoice.tax_amount).label('total_tax')
    ).join(
        PurchaseInvoice, Supplier.id == PurchaseInvoice.supplier_id
    ).filter(
        PurchaseInvoice.status != 'cancelled'
    )

    if start_date:
        query = query.filter(PurchaseInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(PurchaseInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if supplier_id:
        query = query.filter(Supplier.id == supplier_id)

    query = query.group_by(Supplier.id, Supplier.name, Supplier.code)
    query = query.order_by(func.sum(PurchaseInvoice.total_amount).desc())

    suppliers_data = query.all()

    total_amount = sum(s.total_amount or 0 for s in suppliers_data)
    total_invoices = sum(s.invoice_count or 0 for s in suppliers_data)
    total_tax = sum(s.total_tax or 0 for s in suppliers_data)

    all_suppliers = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()

    from app.models import Company
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'SAR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'ريال')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', 'ر.س')

    return render_template('reports/purchases_by_supplier.html',
                         suppliers_data=suppliers_data,
                         total_amount=total_amount,
                         total_invoices=total_invoices,
                         total_tax=total_tax,
                         all_suppliers=all_suppliers,
                         start_date=start_date,
                         end_date=end_date,
                         selected_supplier_id=supplier_id,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)


@bp.route('/purchases-monthly')
@login_required
@permission_required('reports.purchases')
def purchases_monthly():
    """Monthly purchases report"""
    year = request.args.get('year', type=int)
    if not year:
        year = datetime.now().year

    # Get all invoices for the year
    start_date = datetime(year, 1, 1).date()
    end_date = datetime(year, 12, 31).date()

    invoices = PurchaseInvoice.query.filter(
        PurchaseInvoice.status != 'cancelled',
        PurchaseInvoice.invoice_date >= start_date,
        PurchaseInvoice.invoice_date <= end_date
    ).all()

    # Group by month
    monthly_dict = {}
    for invoice in invoices:
        month_num = invoice.invoice_date.month
        if month_num not in monthly_dict:
            monthly_dict[month_num] = {
                'invoice_count': 0,
                'total_amount': 0,
                'total_tax': 0
            }
        monthly_dict[month_num]['invoice_count'] += 1
        monthly_dict[month_num]['total_amount'] += invoice.total_amount
        monthly_dict[month_num]['total_tax'] += invoice.tax_amount

    # Create complete 12-month data
    months_data = []
    month_names = [
        'يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
        'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'
    ]

    for month_num in range(1, 13):
        if month_num in monthly_dict:
            data = monthly_dict[month_num]
            months_data.append({
                'month_num': month_num,
                'month_name': month_names[month_num - 1],
                'invoice_count': data['invoice_count'],
                'total_amount': data['total_amount'],
                'total_tax': data['total_tax']
            })
        else:
            months_data.append({
                'month_num': month_num,
                'month_name': month_names[month_num - 1],
                'invoice_count': 0,
                'total_amount': 0,
                'total_tax': 0
            })

    # Calculate totals
    total_invoices = sum(m['invoice_count'] for m in months_data)
    total_purchases = sum(m['total_amount'] for m in months_data)
    total_tax = sum(m['total_tax'] for m in months_data)

    # Get currency settings
    currency_code = current_app.config.get('CURRENCY', 'EUR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'Euro')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    # Get available years
    all_invoices = PurchaseInvoice.query.filter(
        PurchaseInvoice.status != 'cancelled'
    ).all()

    years_set = set()
    for inv in all_invoices:
        if inv.invoice_date:
            years_set.add(inv.invoice_date.year)

    available_years = sorted(list(years_set), reverse=True)
    if not available_years:
        available_years = [datetime.now().year]

    return render_template('reports/purchases_monthly.html',
                         months_data=months_data,
                         total_invoices=total_invoices,
                         total_purchases=total_purchases,
                         total_tax=total_tax,
                         year=year,
                         available_years=available_years,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)


# ==================== Customer Reports ====================

@bp.route('/customers')
@login_required
@permission_required('reports.sales')
def customers_list():
    """Customers list report"""
    customers = Customer.query.filter_by(is_active=True).all()

    total_customers = len(customers)
    total_balance = sum(c.current_balance or 0 for c in customers)
    customers_with_sales = sum(1 for c in customers if c.invoices)

    from app.models import Company
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'EUR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'Euro')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    return render_template('reports/customers.html',
                         customers=customers,
                         total_customers=total_customers,
                         total_balance=total_balance,
                         customers_with_sales=customers_with_sales,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)


@bp.route('/customers/top')
@login_required
@permission_required('reports.sales')
def customers_top():
    """Best customers report - ranked by sales volume"""
    customers_data = []
    customers = Customer.query.filter_by(is_active=True).all()

    for customer in customers:
        total_sales = db.session.query(func.sum(SalesInvoice.total_amount))\
            .filter(SalesInvoice.customer_id == customer.id)\
            .filter(SalesInvoice.status == 'confirmed')\
            .scalar() or 0

        invoice_count = SalesInvoice.query.filter_by(
            customer_id=customer.id,
            status='confirmed'
        ).count()

        if total_sales > 0:
            customers_data.append({
                'customer': customer,
                'total_sales': total_sales,
                'invoice_count': invoice_count,
                'average_sale': total_sales / invoice_count if invoice_count > 0 else 0
            })

    customers_data.sort(key=lambda x: x['total_sales'], reverse=True)

    total_sales = sum(c['total_sales'] for c in customers_data)
    total_invoices = sum(c['invoice_count'] for c in customers_data)

    from app.models import Company
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'EUR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'Euro')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    return render_template('reports/customers_top.html',
                         customers_data=customers_data,
                         total_sales=total_sales,
                         total_invoices=total_invoices,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)


@bp.route('/customers/balances')
@login_required
@permission_required('reports.sales')
def customers_balances():
    """Customer balances report"""
    customers = Customer.query.filter_by(is_active=True).all()

    total_debit = sum(c.current_balance for c in customers if (c.current_balance or 0) > 0)
    total_credit = abs(sum(c.current_balance for c in customers if (c.current_balance or 0) < 0))
    net_balance = sum(c.current_balance or 0 for c in customers)

    from app.models import Company
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'EUR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'Euro')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    return render_template('reports/customers_balances.html',
                         customers=customers,
                         total_debit=total_debit,
                         total_credit=total_credit,
                         net_balance=net_balance,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)


@bp.route('/customers/history/<int:customer_id>')
@login_required
@permission_required('reports.sales')
def customers_history(customer_id):
    """Customer history report - shows all transactions for a specific customer"""
    customer = Customer.query.get_or_404(customer_id)

    invoices = SalesInvoice.query.filter_by(customer_id=customer_id)\
        .order_by(SalesInvoice.invoice_date.desc()).all()

    total_invoices = len(invoices)
    total_sales = sum(inv.total_amount for inv in invoices if inv.status == 'confirmed')
    total_paid = sum(inv.paid_amount or 0 for inv in invoices)
    total_remaining = sum(inv.remaining_amount or 0 for inv in invoices)

    from app.models import Company
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'EUR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'Euro')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    return render_template('reports/customers_history.html',
                         customer=customer,
                         invoices=invoices,
                         total_invoices=total_invoices,
                         total_sales=total_sales,
                         total_paid=total_paid,
                         total_remaining=total_remaining,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)


# ==================== Supplier Reports ====================

@bp.route('/suppliers')
@login_required
@permission_required('reports.purchases')
def suppliers_list():
    """Suppliers list report"""
    suppliers = Supplier.query.filter_by(is_active=True).all()

    # Calculate totals
    total_suppliers = len(suppliers)
    total_balance = sum(s.current_balance or 0 for s in suppliers)

    # Count suppliers with purchases
    suppliers_with_purchases = 0
    for supplier in suppliers:
        if supplier.invoices:
            suppliers_with_purchases += 1

    # Get currency settings
    from app.models import Company
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'EUR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'Euro')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    return render_template('reports/suppliers.html',
                         suppliers=suppliers,
                         total_suppliers=total_suppliers,
                         total_balance=total_balance,
                         suppliers_with_purchases=suppliers_with_purchases,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)


@bp.route('/suppliers/top')
@login_required
@permission_required('reports.purchases')
def suppliers_top():
    """Best suppliers report - ranked by purchase volume"""
    # Get all suppliers with their purchase totals
    suppliers_data = []

    suppliers = Supplier.query.filter_by(is_active=True).all()

    for supplier in suppliers:
        # Calculate total purchases for this supplier
        total_purchases = db.session.query(func.sum(PurchaseInvoice.total_amount))\
            .filter(PurchaseInvoice.supplier_id == supplier.id)\
            .filter(PurchaseInvoice.status == 'confirmed')\
            .scalar() or 0

        # Count invoices
        invoice_count = PurchaseInvoice.query.filter_by(
            supplier_id=supplier.id,
            status='confirmed'
        ).count()

        if total_purchases > 0:
            suppliers_data.append({
                'supplier': supplier,
                'total_purchases': total_purchases,
                'invoice_count': invoice_count,
                'average_purchase': total_purchases / invoice_count if invoice_count > 0 else 0
            })

    # Sort by total purchases (descending)
    suppliers_data.sort(key=lambda x: x['total_purchases'], reverse=True)

    # Calculate totals
    total_purchases = sum(s['total_purchases'] for s in suppliers_data)
    total_invoices = sum(s['invoice_count'] for s in suppliers_data)

    # Get currency settings
    from app.models import Company
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'EUR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'Euro')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    return render_template('reports/suppliers_top.html',
                         suppliers_data=suppliers_data,
                         total_purchases=total_purchases,
                         total_invoices=total_invoices,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)


@bp.route('/suppliers/balances')
@login_required
@permission_required('reports.purchases')
def suppliers_balances():
    """Supplier balances report"""
    suppliers = Supplier.query.filter_by(is_active=True).all()

    # Calculate totals
    total_debit = sum(s.current_balance for s in suppliers if s.current_balance > 0)
    total_credit = abs(sum(s.current_balance for s in suppliers if s.current_balance < 0))
    net_balance = sum(s.current_balance or 0 for s in suppliers)

    # Get currency settings
    from app.models import Company
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'EUR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'Euro')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    return render_template('reports/suppliers_balances.html',
                         suppliers=suppliers,
                         total_debit=total_debit,
                         total_credit=total_credit,
                         net_balance=net_balance,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)


@bp.route('/cash-flow')
@login_required
@permission_required('reports.view')
def cash_flow():
    """Cash Flow Report - تقرير التدفق النقدي"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Default to current month
    today = datetime.today()
    if not start_date:
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = today.strftime('%Y-%m-%d')

    start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()

    # --- Operating Activities: Cash Inflows ---
    # 1. Cash received from customers (paid sales invoices)
    sales_invoices = SalesInvoice.query.filter(
        SalesInvoice.status == 'confirmed',
        SalesInvoice.invoice_date >= start_dt,
        SalesInvoice.invoice_date <= end_dt
    ).all()
    cash_from_customers = sum(inv.paid_amount or 0 for inv in sales_invoices)

    # 2. Receipts from Payment model (customer receipts)
    customer_receipts = Payment.query.filter(
        Payment.payment_type == 'receipt',
        Payment.party_type == 'customer',
        Payment.status == 'posted',
        Payment.payment_date >= start_dt,
        Payment.payment_date <= end_dt
    ).all()
    cash_receipts_payments = sum(p.amount for p in customer_receipts)

    total_cash_inflow = cash_from_customers + cash_receipts_payments

    # --- Operating Activities: Cash Outflows ---
    # 1. Cash paid to suppliers (paid purchase invoices)
    purchase_invoices = PurchaseInvoice.query.filter(
        PurchaseInvoice.status == 'confirmed',
        PurchaseInvoice.invoice_date >= start_dt,
        PurchaseInvoice.invoice_date <= end_dt
    ).all()
    cash_to_suppliers = sum(inv.paid_amount or 0 for inv in purchase_invoices)

    # 2. Payments to suppliers from Payment model
    supplier_payments = Payment.query.filter(
        Payment.payment_type == 'payment',
        Payment.party_type == 'supplier',
        Payment.status == 'posted',
        Payment.payment_date >= start_dt,
        Payment.payment_date <= end_dt
    ).all()
    cash_payments_payments = sum(p.amount for p in supplier_payments)

    # 3. Expenses paid
    expenses = Expense.query.filter(
        Expense.status == 'posted',
        Expense.expense_date >= start_dt,
        Expense.expense_date <= end_dt
    ).all()
    cash_for_expenses = sum(e.amount for e in expenses)

    total_cash_outflow = cash_to_suppliers + cash_payments_payments + cash_for_expenses

    # Net operating cash flow
    net_operating = total_cash_inflow - total_cash_outflow

    # Expense breakdown by category
    expense_by_category = {}
    for e in expenses:
        cat = e.expense_category or 'other'
        expense_by_category[cat] = expense_by_category.get(cat, 0) + e.amount

    # Currency settings
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'EUR')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    return render_template('reports/cash_flow.html',
                           start_date=start_date,
                           end_date=end_date,
                           cash_from_customers=cash_from_customers,
                           cash_receipts_payments=cash_receipts_payments,
                           total_cash_inflow=total_cash_inflow,
                           cash_to_suppliers=cash_to_suppliers,
                           cash_payments_payments=cash_payments_payments,
                           cash_for_expenses=cash_for_expenses,
                           expense_by_category=expense_by_category,
                           total_cash_outflow=total_cash_outflow,
                           net_operating=net_operating,
                           currency_symbol=currency_symbol)


@bp.route('/accounts-receivable')
@login_required
@permission_required('reports.sales')
def accounts_receivable():
    """Accounts Receivable Report - تقرير حسابات القبض"""
    customer_id = request.args.get('customer_id', type=int)
    status_filter = request.args.get('status', 'unpaid')  # unpaid, partial, all

    today = datetime.today().date()

    # Base query: confirmed invoices with remaining balance
    query = SalesInvoice.query.filter(
        SalesInvoice.status == 'confirmed',
        SalesInvoice.remaining_amount > 0
    )

    if customer_id:
        query = query.filter(SalesInvoice.customer_id == customer_id)

    if status_filter == 'unpaid':
        query = query.filter(SalesInvoice.payment_status == 'unpaid')
    elif status_filter == 'partial':
        query = query.filter(SalesInvoice.payment_status == 'partial')
    # 'all' => no extra filter

    invoices = query.order_by(SalesInvoice.invoice_date.asc()).all()

    # Enrich with aging info
    invoice_data = []
    for inv in invoices:
        days_overdue = (today - inv.invoice_date).days
        if days_overdue <= 30:
            aging_bucket = 'current'
        elif days_overdue <= 60:
            aging_bucket = '31_60'
        elif days_overdue <= 90:
            aging_bucket = '61_90'
        elif days_overdue <= 120:
            aging_bucket = '91_120'
        else:
            aging_bucket = 'over_120'
        invoice_data.append({
            'invoice': inv,
            'days_overdue': days_overdue,
            'aging_bucket': aging_bucket
        })

    # Aging totals
    aging_totals = {
        'current': sum(d['invoice'].remaining_amount for d in invoice_data if d['aging_bucket'] == 'current'),
        '31_60':   sum(d['invoice'].remaining_amount for d in invoice_data if d['aging_bucket'] == '31_60'),
        '61_90':   sum(d['invoice'].remaining_amount for d in invoice_data if d['aging_bucket'] == '61_90'),
        '91_120':  sum(d['invoice'].remaining_amount for d in invoice_data if d['aging_bucket'] == '91_120'),
        'over_120':sum(d['invoice'].remaining_amount for d in invoice_data if d['aging_bucket'] == 'over_120'),
    }

    # Summary
    total_invoices = len(invoice_data)
    total_receivable = sum(d['invoice'].remaining_amount for d in invoice_data)
    total_original = sum(d['invoice'].total_amount for d in invoice_data)
    total_paid = sum(d['invoice'].paid_amount for d in invoice_data)
    unique_customers = len(set(d['invoice'].customer_id for d in invoice_data))

    # All active customers for filter dropdown
    customers = Customer.query.filter_by(is_active=True).order_by(Customer.name).all()

    # Currency settings
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'EUR')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    return render_template('reports/accounts_receivable.html',
                           invoice_data=invoice_data,
                           aging_totals=aging_totals,
                           total_invoices=total_invoices,
                           total_receivable=total_receivable,
                           total_original=total_original,
                           total_paid=total_paid,
                           unique_customers=unique_customers,
                           customers=customers,
                           customer_id=customer_id,
                           status_filter=status_filter,
                           currency_symbol=currency_symbol,
                           today=today)


@bp.route('/suppliers/history/<int:supplier_id>')
@login_required
@permission_required('reports.purchases')
def suppliers_history(supplier_id):
    """Supplier history report - shows all transactions for a specific supplier"""
    supplier = Supplier.query.get_or_404(supplier_id)

    # Get all purchase invoices for this supplier
    invoices = PurchaseInvoice.query.filter_by(supplier_id=supplier_id)\
        .order_by(PurchaseInvoice.invoice_date.desc()).all()

    # Calculate totals
    total_invoices = len(invoices)
    total_purchases = sum(inv.total_amount for inv in invoices if inv.status == 'confirmed')
    total_paid = sum(inv.paid_amount or 0 for inv in invoices)
    total_remaining = sum(inv.remaining_amount or 0 for inv in invoices)

    # Get currency settings
    from app.models import Company
    company = Company.query.first()
    currency_code = company.currency if company else current_app.config.get('DEFAULT_CURRENCY', 'EUR')
    currency_name = current_app.config['CURRENCIES'].get(currency_code, {}).get('name', 'Euro')
    currency_symbol = current_app.config['CURRENCIES'].get(currency_code, {}).get('symbol', '€')

    return render_template('reports/suppliers_history.html',
                         supplier=supplier,
                         invoices=invoices,
                         total_invoices=total_invoices,
                         total_purchases=total_purchases,
                         total_paid=total_paid,
                         total_remaining=total_remaining,
                         currency_code=currency_code,
                         currency_name=currency_name,
                         currency_symbol=currency_symbol)

# ==================== Excel Export Routes ====================

@bp.route('/export-excel/sales')
@login_required
@permission_required('reports.sales')
def export_excel_sales():
    """Export sales report to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from flask import flash
        flash('openpyxl غير مثبت', 'danger')
        return redirect(url_for('reports.sales_report'))

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = SalesInvoice.query.filter_by(status='confirmed')
    if start_date:
        query = query.filter(SalesInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(SalesInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    invoices = query.order_by(SalesInvoice.invoice_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales Report'
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color='1a6b3c', end_color='1a6b3c', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = ['Invoice #', 'Date', 'Customer', 'Total', 'Paid', 'Remaining', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, inv in enumerate(invoices, 2):
        ws.cell(row=row, column=1, value=inv.invoice_number)
        ws.cell(row=row, column=2, value=str(inv.invoice_date))
        ws.cell(row=row, column=3, value=inv.customer.name if inv.customer else '')
        ws.cell(row=row, column=4, value=inv.total_amount)
        ws.cell(row=row, column=5, value=inv.paid_amount)
        ws.cell(row=row, column=6, value=inv.remaining_amount)
        ws.cell(row=row, column=7, value=inv.payment_status)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'sales_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/export-excel/purchases')
@login_required
@permission_required('reports.purchases')
def export_excel_purchases():
    """Export purchases report to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from flask import flash
        flash('openpyxl غير مثبت', 'danger')
        return redirect(url_for('reports.purchases_report'))

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = PurchaseInvoice.query.filter(PurchaseInvoice.status != 'cancelled')
    if start_date:
        query = query.filter(PurchaseInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(PurchaseInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    invoices = query.order_by(PurchaseInvoice.invoice_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Purchases Report'
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color='154360', end_color='154360', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = ['Invoice #', 'Date', 'Supplier', 'Total', 'Paid', 'Remaining', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, inv in enumerate(invoices, 2):
        ws.cell(row=row, column=1, value=inv.invoice_number)
        ws.cell(row=row, column=2, value=str(inv.invoice_date))
        ws.cell(row=row, column=3, value=inv.supplier.name if inv.supplier else '')
        ws.cell(row=row, column=4, value=inv.total_amount)
        ws.cell(row=row, column=5, value=inv.paid_amount)
        ws.cell(row=row, column=6, value=inv.remaining_amount)
        ws.cell(row=row, column=7, value=inv.payment_status)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'purchases_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/export-excel/cash-flow')
@login_required
@permission_required('reports.view')
def export_excel_cash_flow():
    """Export cash flow report to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from flask import flash
        flash('openpyxl غير مثبت', 'danger')
        return redirect(url_for('reports.cash_flow'))

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Reuse same logic as cash_flow route
    today = datetime.today().date()
    sd = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else today.replace(day=1)
    ed = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else today

    sales = SalesInvoice.query.filter(
        SalesInvoice.status == 'confirmed',
        SalesInvoice.invoice_date >= sd,
        SalesInvoice.invoice_date <= ed
    ).all()

    purchases = PurchaseInvoice.query.filter(
        PurchaseInvoice.status == 'confirmed',
        PurchaseInvoice.invoice_date >= sd,
        PurchaseInvoice.invoice_date <= ed
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cash Flow'
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color='0077b6', end_color='0077b6', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    ws.cell(row=1, column=1, value='Cash Flow Report').font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f'Period: {sd} to {ed}')

    ws.cell(row=4, column=1, value='Type')
    ws.cell(row=4, column=2, value='Invoice #')
    ws.cell(row=4, column=3, value='Date')
    ws.cell(row=4, column=4, value='Party')
    ws.cell(row=4, column=5, value='Amount')
    for col in range(1, 6):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    row = 5
    for inv in sales:
        ws.cell(row=row, column=1, value='Inflow (Sales)')
        ws.cell(row=row, column=2, value=inv.invoice_number)
        ws.cell(row=row, column=3, value=str(inv.invoice_date))
        ws.cell(row=row, column=4, value=inv.customer.name if inv.customer else '')
        ws.cell(row=row, column=5, value=inv.total_amount)
        row += 1

    for inv in purchases:
        ws.cell(row=row, column=1, value='Outflow (Purchases)')
        ws.cell(row=row, column=2, value=inv.invoice_number)
        ws.cell(row=row, column=3, value=str(inv.invoice_date))
        ws.cell(row=row, column=4, value=inv.supplier.name if inv.supplier else '')
        ws.cell(row=row, column=5, value=-inv.total_amount)
        row += 1

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'cash_flow_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/export-excel/accounts-receivable')
@login_required
@permission_required('reports.sales')
def export_excel_accounts_receivable():
    """Export accounts receivable report to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from flask import flash
        flash('openpyxl غير مثبت', 'danger')
        return redirect(url_for('reports.accounts_receivable'))

    customer_id = request.args.get('customer_id', type=int)
    status_filter = request.args.get('status', 'unpaid')
    today = datetime.today().date()

    query = SalesInvoice.query.filter(
        SalesInvoice.status == 'confirmed',
        SalesInvoice.remaining_amount > 0
    )
    if customer_id:
        query = query.filter(SalesInvoice.customer_id == customer_id)
    if status_filter == 'unpaid':
        query = query.filter(SalesInvoice.payment_status == 'unpaid')
    elif status_filter == 'partial':
        query = query.filter(SalesInvoice.payment_status == 'partial')
    invoices = query.order_by(SalesInvoice.invoice_date.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Accounts Receivable'
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color='7b3f00', end_color='7b3f00', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = ['Invoice #', 'Date', 'Customer', 'Total', 'Paid', 'Remaining', 'Days Overdue', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, inv in enumerate(invoices, 2):
        days = (today - inv.invoice_date).days
        ws.cell(row=row, column=1, value=inv.invoice_number)
        ws.cell(row=row, column=2, value=str(inv.invoice_date))
        ws.cell(row=row, column=3, value=inv.customer.name if inv.customer else '')
        ws.cell(row=row, column=4, value=inv.total_amount)
        ws.cell(row=row, column=5, value=inv.paid_amount or 0)
        ws.cell(row=row, column=6, value=inv.remaining_amount)
        ws.cell(row=row, column=7, value=days)
        ws.cell(row=row, column=8, value=inv.payment_status)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'accounts_receivable_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ==================== PDF Export Routes ====================

def _arabic(text):
    """Reshape and reorder Arabic text for proper PDF rendering."""
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)
    except Exception:
        return str(text)


def _make_pdf_styles():
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    styles = getSampleStyleSheet()
    # Try to register Arabic-capable font
    font_paths = [
        r'C:\Windows\Fonts\arial.ttf',
        r'C:\Windows\Fonts\tahoma.ttf',
        r'C:\Windows\Fonts\calibri.ttf',
    ]
    arabic_font = 'Helvetica'
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                font_name = os.path.splitext(os.path.basename(fp))[0].capitalize()
                pdfmetrics.registerFont(TTFont(font_name, fp))
                arabic_font = font_name
                break
            except Exception:
                continue

    title_style = ParagraphStyle('ArabicTitle', fontName=arabic_font,
                                  fontSize=16, alignment=TA_CENTER, spaceAfter=12)
    cell_style = ParagraphStyle('ArabicCell', fontName=arabic_font,
                                 fontSize=9, alignment=TA_RIGHT)
    header_style = ParagraphStyle('ArabicHeader', fontName=arabic_font,
                                   fontSize=9, alignment=TA_CENTER, textColor='white')
    return arabic_font, title_style, cell_style, header_style


@bp.route('/export-pdf/sales')
@login_required
@permission_required('reports.sales')
def export_pdf_sales():
    """Export sales report to PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = SalesInvoice.query.filter_by(status='confirmed')
    if start_date:
        query = query.filter(SalesInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(SalesInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    invoices = query.order_by(SalesInvoice.invoice_date.desc()).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm,
                             topMargin=1.5*cm, bottomMargin=1*cm)
    arabic_font, title_style, cell_style, header_style = _make_pdf_styles()

    elements = []
    elements.append(Paragraph(_arabic('تقرير المبيعات - Sales Report'), title_style))
    if start_date or end_date:
        period = f"{start_date or ''} - {end_date or ''}"
        elements.append(Paragraph(_arabic(f'الفترة: {period}'), cell_style))
    elements.append(Spacer(1, 0.3*cm))

    header_color = colors.HexColor('#1a6b3c')
    headers = ['Invoice #', 'Date', 'Customer', 'Total', 'Paid', 'Remaining', 'Status']
    data = [headers]
    for inv in invoices:
        data.append([
            inv.invoice_number or '',
            str(inv.invoice_date) if inv.invoice_date else '',
            _arabic(inv.customer.name if inv.customer else ''),
            f"{inv.total_amount:,.2f}",
            f"{(inv.paid_amount or 0):,.2f}",
            f"{(inv.remaining_amount or 0):,.2f}",
            inv.payment_status or '',
        ])

    col_widths = [3*cm, 2.5*cm, 6*cm, 3*cm, 3*cm, 3*cm, 2.5*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), header_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,-1), arabic_font),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f7f0')]),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    filename = f'sales_report_{datetime.now().strftime("%Y%m%d")}.pdf'
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')


@bp.route('/export-pdf/purchases')
@login_required
@permission_required('reports.purchases')
def export_pdf_purchases():
    """Export purchases report to PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = PurchaseInvoice.query.filter(PurchaseInvoice.status != 'cancelled')
    if start_date:
        query = query.filter(PurchaseInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(PurchaseInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    invoices = query.order_by(PurchaseInvoice.invoice_date.desc()).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm,
                             topMargin=1.5*cm, bottomMargin=1*cm)
    arabic_font, title_style, cell_style, header_style = _make_pdf_styles()

    elements = []
    elements.append(Paragraph(_arabic('تقرير المشتريات - Purchases Report'), title_style))
    if start_date or end_date:
        period = f"{start_date or ''} - {end_date or ''}"
        elements.append(Paragraph(_arabic(f'الفترة: {period}'), cell_style))
    elements.append(Spacer(1, 0.3*cm))

    header_color = colors.HexColor('#154360')
    headers = ['Invoice #', 'Date', 'Supplier', 'Total', 'Paid', 'Remaining', 'Status']
    data = [headers]
    for inv in invoices:
        data.append([
            inv.invoice_number or '',
            str(inv.invoice_date) if inv.invoice_date else '',
            _arabic(inv.supplier.name if inv.supplier else ''),
            f"{inv.total_amount:,.2f}",
            f"{(inv.paid_amount or 0):,.2f}",
            f"{(inv.remaining_amount or 0):,.2f}",
            inv.payment_status or '',
        ])

    col_widths = [3*cm, 2.5*cm, 6*cm, 3*cm, 3*cm, 3*cm, 2.5*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), header_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,-1), arabic_font),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#eaf0fb')]),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    filename = f'purchases_report_{datetime.now().strftime("%Y%m%d")}.pdf'
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')


@bp.route('/export-pdf/cash-flow')
@login_required
@permission_required('reports.view')
def export_pdf_cash_flow():
    """Export cash flow report to PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    today = datetime.today().date()
    sd = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else today.replace(day=1)
    ed = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else today

    sales = SalesInvoice.query.filter(
        SalesInvoice.status == 'confirmed',
        SalesInvoice.invoice_date >= sd,
        SalesInvoice.invoice_date <= ed
    ).all()
    purchases = PurchaseInvoice.query.filter(
        PurchaseInvoice.status == 'confirmed',
        PurchaseInvoice.invoice_date >= sd,
        PurchaseInvoice.invoice_date <= ed
    ).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm,
                             topMargin=1.5*cm, bottomMargin=1*cm)
    arabic_font, title_style, cell_style, header_style = _make_pdf_styles()

    elements = []
    elements.append(Paragraph(_arabic('تقرير التدفق النقدي - Cash Flow Report'), title_style))
    elements.append(Paragraph(_arabic(f'الفترة: {sd} - {ed}'), cell_style))
    elements.append(Spacer(1, 0.3*cm))

    header_color = colors.HexColor('#0077b6')
    headers = ['Type', 'Invoice #', 'Date', 'Party', 'Amount']
    data = [headers]
    for inv in sales:
        data.append(['Inflow (Sales)', inv.invoice_number or '',
                     str(inv.invoice_date), _arabic(inv.customer.name if inv.customer else ''),
                     f"{inv.total_amount:,.2f}"])
    for inv in purchases:
        data.append(['Outflow (Purchases)', inv.invoice_number or '',
                     str(inv.invoice_date), _arabic(inv.supplier.name if inv.supplier else ''),
                     f"-{inv.total_amount:,.2f}"])

    col_widths = [4*cm, 3*cm, 3*cm, 8*cm, 3.5*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), header_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,-1), arabic_font),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#e8f4f8')]),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    filename = f'cash_flow_{datetime.now().strftime("%Y%m%d")}.pdf'
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')


@bp.route('/export-pdf/accounts-receivable')
@login_required
@permission_required('reports.sales')
def export_pdf_accounts_receivable():
    """Export accounts receivable report to PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    customer_id = request.args.get('customer_id', type=int)
    status_filter = request.args.get('status', 'unpaid')
    today = datetime.today().date()

    query = SalesInvoice.query.filter(
        SalesInvoice.status == 'confirmed',
        SalesInvoice.remaining_amount > 0
    )
    if customer_id:
        query = query.filter(SalesInvoice.customer_id == customer_id)
    if status_filter == 'unpaid':
        query = query.filter(SalesInvoice.payment_status == 'unpaid')
    elif status_filter == 'partial':
        query = query.filter(SalesInvoice.payment_status == 'partial')
    invoices = query.order_by(SalesInvoice.invoice_date.asc()).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm,
                             topMargin=1.5*cm, bottomMargin=1*cm)
    arabic_font, title_style, cell_style, header_style = _make_pdf_styles()

    elements = []
    elements.append(Paragraph(_arabic('تقرير حسابات القبض - Accounts Receivable'), title_style))
    elements.append(Spacer(1, 0.3*cm))

    header_color = colors.HexColor('#7b3f00')
    headers = ['Invoice #', 'Date', 'Customer', 'Total', 'Paid', 'Remaining', 'Days', 'Status']
    data = [headers]
    for inv in invoices:
        days = (today - inv.invoice_date).days if inv.invoice_date else 0
        data.append([
            inv.invoice_number or '',
            str(inv.invoice_date) if inv.invoice_date else '',
            _arabic(inv.customer.name if inv.customer else ''),
            f"{inv.total_amount:,.2f}",
            f"{(inv.paid_amount or 0):,.2f}",
            f"{(inv.remaining_amount or 0):,.2f}",
            str(days),
            inv.payment_status or '',
        ])

    col_widths = [2.8*cm, 2.3*cm, 5.5*cm, 2.8*cm, 2.8*cm, 2.8*cm, 1.8*cm, 2.2*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), header_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,-1), arabic_font),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#fdf3e7')]),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    filename = f'accounts_receivable_{datetime.now().strftime("%Y%m%d")}.pdf'
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')



# ==================== PDF Helper ====================

def _make_pdf_table(title, headers, rows, col_widths=None):
    """Build a landscape A4 PDF with a title row and data table. Returns BytesIO."""
    import io
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=1.5*cm, rightMargin=1.5*cm)
    arabic_font, title_style, cell_style, header_style = _make_pdf_styles()

    elements = [Paragraph(_arabic(title), title_style), Spacer(1, 0.4*cm)]

    header_row = [Paragraph(_arabic(h), header_style) for h in headers]
    data_rows = []
    for row in rows:
        data_rows.append([Paragraph(_arabic(str(c)), cell_style) for c in row])

    page_w = landscape(A4)[0] - 3*cm
    if not col_widths:
        col_widths = [page_w / len(headers)] * len(headers)

    t = Table([header_row] + data_rows, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return buf


# ==================== Additional PDF Export Routes ====================

@bp.route('/export-pdf/sales-by-product')
@login_required
@permission_required('reports.sales')
def export_pdf_sales_by_product():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = db.session.query(
        Product.name, Product.code,
        func.sum(SalesInvoiceItem.quantity).label('total_qty'),
        func.sum(SalesInvoiceItem.total).label('total_amount')
    ).join(SalesInvoiceItem).join(SalesInvoice).filter(SalesInvoice.status != 'cancelled')
    if start_date:
        query = query.filter(SalesInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(SalesInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    results = query.group_by(Product.id).order_by(func.sum(SalesInvoiceItem.total).desc()).all()

    rows = [[r.name, r.code or '', f'{r.total_qty or 0:.2f}', f'{r.total_amount or 0:.2f}'] for r in results]
    buf = _make_pdf_table('تقرير المبيعات حسب المنتج',
                          ['المنتج', 'الكود', 'الكمية', 'الإجمالي'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'sales_by_product_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/sales-by-customer')
@login_required
@permission_required('reports.sales')
def export_pdf_sales_by_customer():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = db.session.query(
        Customer.name, Customer.code,
        func.count(SalesInvoice.id).label('invoice_count'),
        func.sum(SalesInvoice.total_amount).label('total_amount')
    ).join(SalesInvoice).filter(SalesInvoice.status != 'cancelled')
    if start_date:
        query = query.filter(SalesInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(SalesInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    results = query.group_by(Customer.id).order_by(func.sum(SalesInvoice.total_amount).desc()).all()

    rows = [[r.name, r.code or '', str(r.invoice_count), f'{r.total_amount or 0:.2f}'] for r in results]
    buf = _make_pdf_table('تقرير المبيعات حسب العميل',
                          ['العميل', 'الكود', 'عدد الفواتير', 'الإجمالي'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'sales_by_customer_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/sales-monthly')
@login_required
@permission_required('reports.sales')
def export_pdf_sales_monthly():
    year = request.args.get('year', type=int) or datetime.now().year
    invoices = SalesInvoice.query.filter(
        SalesInvoice.status != 'cancelled',
        SalesInvoice.invoice_date >= datetime(year, 1, 1).date(),
        SalesInvoice.invoice_date <= datetime(year, 12, 31).date()
    ).all()

    monthly = {}
    for inv in invoices:
        m = inv.invoice_date.month
        monthly.setdefault(m, {'count': 0, 'amount': 0})
        monthly[m]['count'] += 1
        monthly[m]['amount'] += inv.total_amount

    month_names = ['يناير','فبراير','مارس','أبريل','مايو','يونيو',
                   'يوليو','أغسطس','سبتمبر','أكتوبر','نوفمبر','ديسمبر']
    rows = []
    for i in range(1, 13):
        d = monthly.get(i, {'count': 0, 'amount': 0})
        rows.append([month_names[i-1], str(d['count']), f"{d['amount']:.2f}"])

    buf = _make_pdf_table(f'تقرير المبيعات الشهري - {year}',
                          ['الشهر', 'عدد الفواتير', 'الإجمالي'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'sales_monthly_{year}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/purchases-by-product')
@login_required
@permission_required('reports.purchases')
def export_pdf_purchases_by_product():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = db.session.query(
        Product.name, Product.code,
        func.sum(PurchaseInvoiceItem.quantity).label('total_quantity'),
        func.sum(PurchaseInvoiceItem.total).label('total_amount'),
        func.count(func.distinct(PurchaseInvoiceItem.invoice_id)).label('invoice_count')
    ).join(PurchaseInvoiceItem, Product.id == PurchaseInvoiceItem.product_id
    ).join(PurchaseInvoice, PurchaseInvoiceItem.invoice_id == PurchaseInvoice.id
    ).filter(PurchaseInvoice.status != 'cancelled')
    if start_date:
        query = query.filter(PurchaseInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(PurchaseInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    results = query.group_by(Product.id, Product.name, Product.code
                             ).order_by(func.sum(PurchaseInvoiceItem.total).desc()).all()

    rows = [[r.name, r.code or '', f'{r.total_quantity or 0:.2f}',
             str(r.invoice_count), f'{r.total_amount or 0:.2f}'] for r in results]
    buf = _make_pdf_table('تقرير المشتريات حسب المنتج',
                          ['المنتج', 'الكود', 'الكمية', 'الفواتير', 'الإجمالي'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'purchases_by_product_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/purchases-by-supplier')
@login_required
@permission_required('reports.purchases')
def export_pdf_purchases_by_supplier():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = db.session.query(
        Supplier.name, Supplier.code,
        func.count(func.distinct(PurchaseInvoice.id)).label('invoice_count'),
        func.sum(PurchaseInvoice.total_amount).label('total_amount')
    ).join(PurchaseInvoice, Supplier.id == PurchaseInvoice.supplier_id
    ).filter(PurchaseInvoice.status != 'cancelled')
    if start_date:
        query = query.filter(PurchaseInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(PurchaseInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    results = query.group_by(Supplier.id, Supplier.name, Supplier.code
                             ).order_by(func.sum(PurchaseInvoice.total_amount).desc()).all()

    rows = [[r.name, r.code or '', str(r.invoice_count), f'{r.total_amount or 0:.2f}'] for r in results]
    buf = _make_pdf_table('تقرير المشتريات حسب المورد',
                          ['المورد', 'الكود', 'عدد الفواتير', 'الإجمالي'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'purchases_by_supplier_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/purchases-monthly')
@login_required
@permission_required('reports.purchases')
def export_pdf_purchases_monthly():
    year = request.args.get('year', type=int) or datetime.now().year
    invoices = PurchaseInvoice.query.filter(
        PurchaseInvoice.status != 'cancelled',
        PurchaseInvoice.invoice_date >= datetime(year, 1, 1).date(),
        PurchaseInvoice.invoice_date <= datetime(year, 12, 31).date()
    ).all()

    monthly = {}
    for inv in invoices:
        m = inv.invoice_date.month
        monthly.setdefault(m, {'count': 0, 'amount': 0})
        monthly[m]['count'] += 1
        monthly[m]['amount'] += inv.total_amount

    month_names = ['يناير','فبراير','مارس','أبريل','مايو','يونيو',
                   'يوليو','أغسطس','سبتمبر','أكتوبر','نوفمبر','ديسمبر']
    rows = []
    for i in range(1, 13):
        d = monthly.get(i, {'count': 0, 'amount': 0})
        rows.append([month_names[i-1], str(d['count']), f"{d['amount']:.2f}"])

    buf = _make_pdf_table(f'تقرير المشتريات الشهري - {year}',
                          ['الشهر', 'عدد الفواتير', 'الإجمالي'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'purchases_monthly_{year}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/inventory')
@login_required
@permission_required('reports.inventory')
def export_pdf_inventory():
    products = Product.query.filter_by(is_active=True, track_inventory=True).all()
    rows = []
    for p in products:
        qty = p.get_stock()
        value = qty * p.cost_price
        rows.append([p.name, p.code or '', f'{qty:.2f}', f'{p.cost_price:.2f}', f'{value:.2f}'])
    buf = _make_pdf_table('تقرير المخزون الحالي',
                          ['المنتج', 'الكود', 'الكمية', 'سعر التكلفة', 'القيمة'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'inventory_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/low-stock')
@login_required
@permission_required('reports.inventory')
def export_pdf_low_stock():
    products = Product.query.filter_by(is_active=True, track_inventory=True).all()
    rows = []
    for p in products:
        qty = p.get_stock()
        if qty <= p.min_stock:
            rows.append([p.name, p.code or '', f'{qty:.2f}',
                         f'{p.min_stock:.2f}', f'{p.min_stock - qty:.2f}'])
    buf = _make_pdf_table('تقرير منتجات تحت الحد الأدنى',
                          ['المنتج', 'الكود', 'المخزون الحالي', 'الحد الأدنى', 'النقص'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'low_stock_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/stock-movement')
@login_required
@permission_required('reports.inventory')
def export_pdf_stock_movement():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = StockMovement.query
    if start_date:
        query = query.filter(StockMovement.created_at >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(StockMovement.created_at <= datetime.strptime(end_date + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    movements = query.order_by(StockMovement.created_at.desc()).all()

    rows = []
    for m in movements:
        date_str = m.created_at.strftime('%Y-%m-%d') if m.created_at else ''
        product_name = m.product.name if m.product else ''
        warehouse_name = m.warehouse.name if m.warehouse else ''
        rows.append([date_str, product_name, warehouse_name,
                     m.movement_type or '', f'{m.quantity:.2f}'])
    buf = _make_pdf_table('تقرير حركة المخزون',
                          ['التاريخ', 'المنتج', 'المستودع', 'النوع', 'الكمية'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'stock_movement_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/profit-loss')
@login_required
@permission_required('reports.financial')
def export_pdf_profit_loss():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    revenue_query = db.session.query(func.sum(SalesInvoice.total_amount)).filter(
        SalesInvoice.status != 'cancelled')
    if start_date:
        revenue_query = revenue_query.filter(
            SalesInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        revenue_query = revenue_query.filter(
            SalesInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    total_revenue = revenue_query.scalar() or 0

    sales_q = SalesInvoice.query.filter(SalesInvoice.status != 'cancelled')
    if start_date:
        sales_q = sales_q.filter(SalesInvoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        sales_q = sales_q.filter(SalesInvoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    total_cogs = sum(item.product.cost_price * item.quantity
                     for inv in sales_q.all() for item in inv.items)
    gross_profit = total_revenue - total_cogs

    rows = [
        ['إجمالي الإيرادات', f'{total_revenue:.2f}'],
        ['تكلفة البضاعة المباعة', f'{total_cogs:.2f}'],
        ['مجمل الربح', f'{gross_profit:.2f}'],
    ]
    buf = _make_pdf_table('قائمة الأرباح والخسائر', ['البيان', 'المبلغ'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'profit_loss_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/customers-list')
@login_required
@permission_required('reports.sales')
def export_pdf_customers_list():
    customers = Customer.query.filter_by(is_active=True).all()
    rows = [[c.name, c.code or '', c.phone or '', f'{c.current_balance or 0:.2f}'] for c in customers]
    buf = _make_pdf_table('قائمة العملاء', ['الاسم', 'الكود', 'الهاتف', 'الرصيد'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'customers_list_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/customers-top')
@login_required
@permission_required('reports.sales')
def export_pdf_customers_top():
    customers = Customer.query.filter_by(is_active=True).all()
    data = []
    for c in customers:
        total = db.session.query(func.sum(SalesInvoice.total_amount)).filter(
            SalesInvoice.customer_id == c.id, SalesInvoice.status == 'confirmed').scalar() or 0
        count = SalesInvoice.query.filter_by(customer_id=c.id, status='confirmed').count()
        if total > 0:
            data.append((c.name, count, total))
    data.sort(key=lambda x: x[2], reverse=True)
    rows = [[name, str(cnt), f'{amt:.2f}'] for name, cnt, amt in data]
    buf = _make_pdf_table('أفضل العملاء', ['العميل', 'عدد الفواتير', 'إجمالي المبيعات'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'customers_top_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/customers-balances')
@login_required
@permission_required('reports.sales')
def export_pdf_customers_balances():
    customers = Customer.query.filter_by(is_active=True).all()
    rows = [[c.name, c.code or '', f'{c.current_balance or 0:.2f}'] for c in customers]
    buf = _make_pdf_table('أرصدة العملاء', ['العميل', 'الكود', 'الرصيد'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'customers_balances_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/suppliers-list')
@login_required
@permission_required('reports.purchases')
def export_pdf_suppliers_list():
    suppliers = Supplier.query.filter_by(is_active=True).all()
    rows = [[s.name, s.code or '', s.phone or '', f'{s.current_balance or 0:.2f}'] for s in suppliers]
    buf = _make_pdf_table('قائمة الموردين', ['الاسم', 'الكود', 'الهاتف', 'الرصيد'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'suppliers_list_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/suppliers-top')
@login_required
@permission_required('reports.purchases')
def export_pdf_suppliers_top():
    suppliers = Supplier.query.filter_by(is_active=True).all()
    data = []
    for s in suppliers:
        total = db.session.query(func.sum(PurchaseInvoice.total_amount)).filter(
            PurchaseInvoice.supplier_id == s.id, PurchaseInvoice.status == 'confirmed').scalar() or 0
        count = PurchaseInvoice.query.filter_by(supplier_id=s.id, status='confirmed').count()
        if total > 0:
            data.append((s.name, count, total))
    data.sort(key=lambda x: x[2], reverse=True)
    rows = [[name, str(cnt), f'{amt:.2f}'] for name, cnt, amt in data]
    buf = _make_pdf_table('أفضل الموردين', ['المورد', 'عدد الفواتير', 'إجمالي المشتريات'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'suppliers_top_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')


@bp.route('/export-pdf/suppliers-balances')
@login_required
@permission_required('reports.purchases')
def export_pdf_suppliers_balances():
    suppliers = Supplier.query.filter_by(is_active=True).all()
    rows = [[s.name, s.code or '', f'{s.current_balance or 0:.2f}'] for s in suppliers]
    buf = _make_pdf_table('أرصدة الموردين', ['المورد', 'الكود', 'الرصيد'], rows)
    return send_file(buf, as_attachment=True,
                     download_name=f'suppliers_balances_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')
